# from django.shortcuts import render, redirect
# from .models import *
# import pandas as pd


# def base(request):
#     return render(request, 'dashboard.html')

# from django.shortcuts import render
# from .models import fund_raising_form

# def dashboard(request):
#     fundraisers = fund_raising_form.objects.all().order_by('-id')[:4]  # latest 4 trending
#     return render(request, 'dashboard.html', {'fundraisers': fundraisers})
 

# # views.py
# from django.shortcuts import render, redirect
# from django.contrib import messages
# from django.contrib.auth.hashers import make_password
# from .models import Signup


# def fund_raising(request):
#     if request.method == 'POST':
#         fund_purpose = request.POST.get('fund_purpose')
#         fund_raise_amount = request.POST.get('fund_raise_amount')
#         fund_title = request.POST.get('fund_title')
#         fund_for = request.POST.get('fund_for')
#         user_qualification = request.POST.get('user_qualification')
#         user_emp_status = request.POST.get('user_emp_status')
#         how_to_know = request.POST.get('how_to_know')


#         raising_fund_info.objects.create(
#             fund_purpose=fund_purpose,
#             fund_raise_amount=fund_raise_amount,
#             fund_title=fund_title,
#             fund_for=fund_for,
#             user_qualification=user_qualification,
#             user_emp_status=user_emp_status,
#             how_to_know=how_to_know
#         )

#         messages.success(request, 'Fund information submitted successfully!')
#         return redirect('dashboard') 

#     return render(request, 'fund_raising.html')



# from django.shortcuts import render, redirect
# from .models import fund_raising_form
# from datetime import date

# def fund_raising_form_view(request):
#     if request.method == "POST":
#         category = request.POST.get("category", "")
#         fund_raising_amount = request.POST.get("fund_raising_amount", "0")
#         try:
#             fund_raising_amount = int(fund_raising_amount)
#         except ValueError:
#             fund_raising_amount = 0

#         common_fields = {
#             "user_info": "John Doe",
#             "user_id": 123,
#             "category": category,
#             "fund_raising_amount": fund_raising_amount,
#             "amount_raised": 0,
#             "status": "Pending",
#             "attachments": request.FILES.get("attachments"),
#             "medical_documents": request.FILES.get("medical_documents"),
#             "start_date": date.today(),
#             "end_date": date.today(),
#             "created_date": date.today(),
#             "created_by": "system",
#             "last_created_date": date.today(),
#             "last_created_by": "system"
#         }

#         # Apply form-specific logic
#         if category == "medical":
#             common_fields.update({
#                 "fund_for": request.POST.get("fund_for", ""),
#                 "patienet_full_name": request.POST.get("patienet_full_name", ""),
#                 "patienet_age": request.POST.get("patienet_age", 0),
#                 "hospital_name": request.POST.get("hospital_name", ""),
#                 "hospital_location": request.POST.get("hospital_location", ""),
#                 "hospitalization_status": request.POST.get("hospitalization_status", ""),
#                 "fund_purpose": request.POST.get("fund_purpose", ""),
#             })

#         elif category == "ngo":
#             common_fields.update({
#                 "fund_title": request.POST.get("fund_title", ""),
#                 "ngo_name": request.POST.get("ngo_name", ""),
#                 "ngo_registration_number": request.POST.get("ngo_registration_number", ""),
#                 "ngo_website_url": request.POST.get("ngo_website_url", ""),
#                 "user_qualification": request.POST.get("user_qualification", ""),
#                 "user_employed_status": request.POST.get("user_employed_status", ""),
#                 "how_to_know": request.POST.get("how_to_know", "")
#             })

#         elif category == "other":
#             common_fields.update({
#                 "fund_title": request.POST.get("fund_title", ""),
#                 "fund_for": request.POST.get("fund_for", ""),
#                 "user_qualification": request.POST.get("user_qualification", ""),
#                 "user_employed_status": request.POST.get("user_employed_status", ""),
#                 "how_to_know": request.POST.get("how_to_know", ""),
#                 "fund_purpose": request.POST.get("fund_purpose", "")
#             })

#         form_data = fund_raising_form(**common_fields)
#         form_data.save()
#         return redirect('dashboard')

#     return render(request, 'fund_raising_form.html')



# # ------------------------------------------------------------------------------------


# from django.shortcuts import render, redirect
# from .models import fund_raising_form
# from datetime import date

# def select_form_category(request):
#     return render(request, 'select_category.html')

# import base64

# from datetime import date
# from django.shortcuts import render, redirect
# from .models import fund_raising_form

# def medical_form(request):
#     if request.method == 'POST':
#         form_data = fund_raising_form(
#             category=request.POST.get("category"),
#             fund_for=request.POST.get("fund_for"),
#             fund_title=request.POST.get("fund_title"),
#             patienet_full_name=request.POST.get("patienet_full_name"),
#             patienet_age=request.POST.get("patienet_age"),
#             hospital_name=request.POST.get("hospital_name"),
#             hospital_location=request.POST.get("hospital_location"),
#             hospitalization_status=request.POST.get("hospitalization_status"),
#             fund_raising_amount=request.POST.get("fund_raising_amount"),
#             fund_purpose=request.POST.get("fund_purpose"),
#             start_date=date.today(),
#             end_date=date.today(),
#             created_date=date.today(),
#             user_id=123,
#             user_info="Unknown",
#             created_by="system",
#         )

#         # Convert image file to base64 string
#         if request.FILES.get("attachments"):
#             image_file = request.FILES.get("attachments")
#             image_data = image_file.read()
#             encoded_image = base64.b64encode(image_data).decode('utf-8')
#             form_data.patienet_image = encoded_image

#         form_data.save()
#         return redirect('dashboard')

#     return render(request, 'medical_form.html')


# from datetime import date
# from django.shortcuts import render, redirect
# from .models import fund_raising_form

# def ngo_form(request):
#     if request.method == 'POST':
#         form_data = fund_raising_form(
#             category=request.POST.get("category"),
#             fund_title=request.POST.get("fund_title", ""),
#             ngo_name=request.POST.get("ngo_name", ""),
#             ngo_registration_number=request.POST.get("ngo_registration_number", ""),
#             ngo_website_url=request.POST.get("ngo_website_url", ""),
#             user_qualification=request.POST.get("user_qualification", ""),
#             user_employed_status=request.POST.get("user_employed_status", ""),
#             how_to_know=request.POST.get("how_to_know", ""),
#             fund_raising_amount=request.POST.get("fund_raising_amount", 0),
#             start_date=date.today(),
#             end_date=date.today(),
#             created_date=date.today(),
#             user_id=123,
#             user_info="Unknown",
#             created_by="system",
#         )
#         form_data.save()
#         return redirect('dashboard')

#     return render(request, 'ngo_form.html')

# import base64
# from datetime import date
# from django.shortcuts import render, redirect
# from .models import fund_raising_form

# def other_form(request):
#     if request.method == 'POST':
#         form_data = fund_raising_form(
#             category=request.POST.get("category"),
#             fund_title=request.POST.get("fund_title", ""),
#             fund_for=request.POST.get("fund_for", ""),
#             fund_raising_amount=request.POST.get("fund_raising_amount", 0),
#             user_qualification=request.POST.get("user_qualification", ""),
#             user_employed_status=request.POST.get("user_employed_status", ""),
#             how_to_know=request.POST.get("how_to_know", ""),
#             fund_purpose=request.POST.get("fund_purpose", ""),
#             start_date=date.today(),
#             end_date=date.today(),
#             created_date=date.today(),
#             user_id=123,
#             user_info="Unknown",
#             created_by="system",
#         )

#         # Convert image file to base64 string if uploaded
#         if request.FILES.get("attachments"):
#             image_file = request.FILES.get("attachments")
#             image_data = image_file.read()
#             encoded_image = base64.b64encode(image_data).decode('utf-8')
#             form_data.patienet_image = encoded_image

#         # Save medical_documents file normally if your model allows it (optional)
#         if request.FILES.get("medical_documents"):
#             form_data.medical_documents = request.FILES.get("medical_documents")

#         form_data.save()
#         return redirect('dashboard')

#     return render(request, 'other_form.html')


# from .models import fund_raising_form
# from django.http import HttpResponse

# def excel(request):
#     data = []
#     if 'show' in request.GET:
#         data = fund_raising_form.objects.all()
#         return render(request, 'excel.html', {'data': data})
#     return HttpResponse("Excel data or file here")

# def patientinfo(request):
#     if request.method == 'POST':
#         request.session['patient_data'] = {
#             'lead_generate_by': request.POST.get('lead_generate_by'),
#             'lead_purpose': request.POST.get('lead_purpose'),
#             'patient_name': request.POST.get('name'),
#             'patient_age':request.POST.get('age'),
#             'patient_case': request.POST.get('case'),
#             'patient_education': request.POST.get('edu'),
#             'patient_occupation': request.POST.get('occu'),
#             'estimate_amount': request.POST.get('amount'),
#             'estimate_date': request.POST.get('date'),
#             'claim_amount': request.POST.get('claim'),
#             'admit_date': request.POST.get('admit'),
#             'patient_address': request.POST.get('address'),
#             'patient_pic_2': request.FILES.get('photo').name if request.FILES.get('photo') else None,
#             'patient_id': request.FILES.get('id').name if request.FILES.get('id') else None,
#         }
#         request.session['patient_files']= {
#             'patient_id': request.FILES.get('id'),
#             'patient_pic_2': request.FILES.get('photo')
#         }               
#         return redirect('guardianinfo') 
#     return render(request, 'patientinfo.html')

# def cards(request):
#     lead = MedicalLead.objects.last()
#     return render(request, 'cards.html', {'lead': lead})
        

# def guardianinfo(request):
#     if request.method == 'POST':
#         request.session['guardian_data'] = {
#             'guardian_name': request.POST.get('name'),
#             'guardian_phone': request.POST.get('phone'),
#         }
#         return redirect('doctorinfo')
#     return render(request, 'guardianinfo.html')

# def doctorinfo(request):
#     if request.method == 'POST':
#         request.session['doctor_Data'] = {
#             'hospital_name': request.POST.get('name'),
#             'doctor_phone': request.POST.get('phone'),
#             'case_type': request.POST.get('case'),
#             'lead_source': request.POST.get('source'),
#         }
        
#         final_submit(request)
        
#         request.session.flush()
        
#         return redirect('success_page')
#     return render(request, 'doctorinfo.html')
    
    
# def final_submit(request):
#         patient_data = request.session.get('patient_data', {})
#         guardian_data = request.session.get('guardian_data', {})
#         doctor_data = request.session.get('doctor_data', {})
#         patient_files = request.session.get('patient_files', {})
        
#         prescription_file = request.FILES.get('prescription'),
#         description_file = request.FILES.get('description'),
        
#         all_data = {
#             **patient_data,
#             **guardian_data,
#             **doctor_data,
#             **patient_files,
#             'prescription_documents': prescription_file,
#             'description_documents':  description_file,
#         }
        
#         MedicalLead.objects.create(**all_data)
        
# def success_page(request):
#             return render(request, 'success.html')
        
# import openpyxl
# from django.http import HttpResponse
# from django.core.files.storage import FileSystemStorage

# def download_excel_page(request):
#     return render(request, 'download_medicalleads.html')
    
# def download_medicalleads_excel(request):
#     # Create workbook and sheet
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "MedicalLeads"

#     # Header row – must match model fields (excluding files/images)
#     headers = [
#         'lead_generate_by', 'lead_purpose', 'patient_name', 'patient_address', 'patient_age',
#         'patient_case', 'claim_amount', 'estimate_amount', 'estimate_date', 'patient_education',
#         'patient_occupation', 'guardian_name', 'guardian_number', 'hospital_name',
#         'doctor_number', 'case_type', 'lead_source'
#     ]
#     ws.append(headers)

#     # Data rows
#     for lead in MedicalLead.objects.all():
#         ws.append([
#             lead.lead_generate_by,
#             lead.lead_purpose,
#             lead.patient_name,
#             lead.patient_address,
#             lead.patient_age,
#             lead.patient_case,
#             lead.claim_amount,
#             lead.estimate_amount,
#             lead.estimate_date,
#             lead.patient_education,
#             lead.patient_occupation,
#             lead.guardian_name,
#             lead.guardian_number,
#             lead.hospital_name,
#             lead.doctor_number,
#             lead.case_type,
#             lead.lead_source
#         ])

#     # Send Excel as downloadable response
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=medicalleads.xlsx'

#     wb.save(response)
#     return response

from django.shortcuts import render

def base(request):
    return render(request, 'dashboard.html') 

def dashboard(request):
    return render(request, 'dashboard.html')

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login as auth_login
# from .models import Student
from .models import MedicalLead
from .models import *
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import make_password
# from .models import Signup
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User



def base(request):
    return render(request, 'dashboard.html')

def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('register')

def dashboard(request):
     return render(request, 'dashboard.html')

# def register(request):
#     if request.method == "POST":
#         full_name = request.POST.get('full_name')
#         email = request.POST.get('email')
#         password = request.POST.get('password')
#         mobile = request.POST.get('mobile')

#         # 🔹 Agar user pehle se hai
#         if User.objects.filter(email=email).exists():
#             return redirect('login')

#         # 🔹 New user → register
#         user = User.objects.create_user(
#             username=email,
#             password=password,
#         )

#         return redirect('login')

#     return render(request, 'register.html')

def register(request):
    if request.method == "POST":
        full_name = request.POST.get('full_name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        mobile = request.POST.get('mobile')

        username = email   # email ko username bana rahe

        User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=full_name
        )

        return redirect('login')

    return render(request, 'register.html')

def signup(request):
    if request.method == "POST":
        full_name = request.POST['full_name']
        email = request.POST['email']
        password = request.POST['password']

        if User.objects.filter(email=email).exists():
            return render(request, 'signup.html', {'error': 'Email already exists'})

        user = User.objects.create_user(
            email=email,
            password=password,
            full_name=full_name
        )

        request.session['user_id'] = user.id
        return redirect('register')

    return render(request, 'register.html')




# def login(request):
#     if request.method == "POST":
#         email = request.POST.get['email']
#         password = request.POST.get['password']

#         user = authenticate(request, email=email, password=password)

#         if user:
#             login(request, user)
#             return redirect('home')

#         return render(request, 'login.html', {'error': 'Invalid credentials'})

#     return render(request, 'login.html') 
from django.contrib.auth import authenticate, login

def login(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        user = authenticate(username=email, password=password)

        if user is not None:
            login(request, user)
            return redirect("home")
        else:
            return render(request, "login.html", {
                "error": "Invalid credentials"
            })

    return render(request, "login.html")


def logout(request):
    logout(request)
    return redirect('login')

def patientinfo(request):
    if request.method == 'POST':
        request.session['patient_data'] = {
            'lead_generate_by': request.POST.get('lead_generate_by'),
            'lead_purpose': request.POST.get('lead_purpose'),
            'patient_name': request.POST.get('name'),
            'patient_age':request.POST.get('age'),
            'patient_case': request.POST.get('case'),
            'patient_education': request.POST.get('edu'),
            'patient_occupation': request.POST.get('occu'),
            'estimate_amount': request.POST.get('amount'),
            'estimate_date': request.POST.get('date'),
            'claim_amount': request.POST.get('claim'),
            'admit_date': request.POST.get('admit'),
            'patient_address': request.POST.get('address'),
            'patient_pic_2': request.FILES.get('photo').name if request.FILES.get('photo') else None,
            'patient_id': request.FILES.get('id').name if request.FILES.get('id') else None,
        }
        request.session['patient_files']= {
            'patient_id': request.FILES.get('id'),
            'patient_pic_2': request.FILES.get('photo')
        }               
        return redirect('guardianinfo') 
    return render(request, 'patientinfo.html')

def cards(request):
    lead = MedicalLead.objects.last()
    return render(request, 'cards.html', {'lead': lead})
        

def guardianinfo(request):
    if request.method == 'POST':
        request.session['guardian_data'] = {
            'guardian_name': request.POST.get('name'),
            'guardian_phone': request.POST.get('phone'),
        }
        return redirect('doctorinfo')
    return render(request, 'guardianinfo.html')

def doctorinfo(request):
    if request.method == 'POST':
        request.session['doctor_data'] = {
            'hospital_name': request.POST.get('name'),
            'doctor_phone': request.POST.get('phone'),
            'case_type': request.POST.get('case'),
            'lead_source': request.POST.get('source'),
        }
        
        final_submit(request)
        
        request.session.flush()
        
        return redirect('success_page')
    return render(request, 'doctorinfo.html')
    
    
def final_submit(request):
        patient_data = request.session.get('patient_data', {})
        guardian_data = request.session.get('guardian_data', {})
        doctor_data = request.session.get('doctor_data', {})
        patient_files = request.session.get('patient_files', {})
        
        prescription_file = request.FILES.get('prescription'),
        description_file = request.FILES.get('description'),
        
        all_data = {
            **patient_data,
            **guardian_data,
            **doctor_data,
            **patient_files,
            'prescription_documents': prescription_file,
            'description_documents':  description_file,
        }
        
        MedicalLead.objects.create(**all_data)
        
def success_page(request):
            return render(request, 'success.html')
        
# def studentinfo(request):
#     if request.method == 'POST':
#         name = request.POST['name']
#         student_class = request.POST['student_class']
#         roll_number = request.POST['roll_number']
#         Student.objects.create(name=name, student_class=student_class, roll_number=roll_number)
#         return redirect('studentinfo')
    
#     students = Student.objects.all()
#     return render(request, 'studentinfo.html', {'students': students})
        
        
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import make_password
        
# def signup(request):
#     if request.method == 'POST':
#         full_name = request.POST.get('full_name')
#         email = request.POST.get('email')
#         password = request.POST.get('password')
#         mobile_number = request.POST.get('mobile')
#         creation_purpose = request.POST.get('creation_purpose')

#         hashed_password = make_password(password)

#         Signup.objects.create(
#             full_name=full_name,
#             email=email,
#             password=hashed_password,
#             mobile_number=mobile_number,
#             creation_purpose=creation_purpose
#         )

#         return redirect('login')

#     return render(request, 'signup.html')



def fund_raising(request):
    if request.method == 'POST':
        fund_purpose = request.POST.get('fund_purpose')
        fund_raise_amount = request.POST.get('fund_raise_amount')
        fund_title = request.POST.get('fund_title')
        fund_for = request.POST.get('fund_for')
        user_qualification = request.POST.get('user_qualification')
        user_emp_status = request.POST.get('user_emp_status')
        how_to_know = request.POST.get('how_to_know')


        raising_fund_info.objects.create(
            fund_purpose=fund_purpose,
            fund_raise_amount=fund_raise_amount,
            fund_title=fund_title,
            fund_for=fund_for,
            user_qualification=user_qualification,
            user_emp_status=user_emp_status,
            how_to_know=how_to_know
        )

        messages.success(request, 'Fund information submitted successfully!')
        return redirect('dashboard') 

    return render(request, 'fund_raising.html')

def medical_form(request):
    if request.method == 'POST':
        form_data = fund_raising_form(
            category=request.POST.get("category"),
            fund_for=request.POST.get("fund_for"),
            fund_title=request.POST.get("fund_title"),
            patienet_full_name=request.POST.get("patienet_full_name"),
            patienet_age=request.POST.get("patienet_age"),
            hospital_name=request.POST.get("hospital_name"),
            hospital_location=request.POST.get("hospital_location"),
            hospitalization_status=request.POST.get("hospitalization_status"),
            fund_raising_amount=request.POST.get("fund_raising_amount"),
            fund_purpose=request.POST.get("fund_purpose"),
            start_date=date.today(),
            end_date=date.today(),
            created_date=date.today(),
            user_id=123,
            user_info="Unknown",
            created_by="system",
        )

        # Convert image file to base64 string
        if request.FILES.get("attachments"):
            image_file = request.FILES.get("attachments")
            image_data = image_file.read()
            encoded_image = base64.b64encode(image_data).decode('utf-8')
            form_data.patienet_image = encoded_image

        form_data.save()
        return redirect('dashboard')

    return render(request, 'medical_form.html')

from django.shortcuts import render, redirect
from datetime import date

def fund_raising_form_view(request):
    if request.method == "POST":
        category = request.POST.get("category", "")
        fund_raising_amount = request.POST.get("fund_raising_amount", "0")
        try:
            fund_raising_amount = int(fund_raising_amount)
        except ValueError:
            fund_raising_amount = 0

        common_fields = {
            "user_info": "John Doe",
            "user_id": 123,
            "category": category,
            "fund_raising_amount": fund_raising_amount,
            "amount_raised": 0,
            "status": "Pending",
            "attachments": request.FILES.get("attachments"),
            "medical_documents": request.FILES.get("medical_documents"),
            "start_date": date.today(),
            "end_date": date.today(),
            "created_date": date.today(),
            "created_by": "system",
            "last_created_date": date.today(),
            "last_created_by": "system"
        }

        # Apply form-specific logic
        if category == "medical":
            common_fields.update({
                "fund_for": request.POST.get("fund_for", ""),
                "patienet_full_name": request.POST.get("patienet_full_name", ""),
                "patienet_age": request.POST.get("patienet_age", 0),
                "hospital_name": request.POST.get("hospital_name", ""),
                "hospital_location": request.POST.get("hospital_location", ""),
                "hospitalization_status": request.POST.get("hospitalization_status", ""),
                "fund_purpose": request.POST.get("fund_purpose", ""),
            })

        elif category == "ngo":
            common_fields.update({
                "fund_title": request.POST.get("fund_title", ""),
                "ngo_name": request.POST.get("ngo_name", ""),
                "ngo_registration_number": request.POST.get("ngo_registration_number", ""),
                "ngo_website_url": request.POST.get("ngo_website_url", ""),
                "user_qualification": request.POST.get("user_qualification", ""),
                "user_employed_status": request.POST.get("user_employed_status", ""),
                "how_to_know": request.POST.get("how_to_know", "")
            })

        elif category == "other":
            common_fields.update({
                "fund_title": request.POST.get("fund_title", ""),
                "fund_for": request.POST.get("fund_for", ""),
                "user_qualification": request.POST.get("user_qualification", ""),
                "user_employed_status": request.POST.get("user_employed_status", ""),
                "how_to_know": request.POST.get("how_to_know", ""),
                "fund_purpose": request.POST.get("fund_purpose", "")
            })

        form_data = fund_raising_form(**common_fields)
        form_data.save()
        return redirect('dashboard')

    return render(request, 'fund_raising_form.html')


from django.shortcuts import render, redirect
from datetime import date

def select_form_category(request):
    return render(request, 'select_category.html')

import base64

from datetime import date
from django.shortcuts import render, redirect

def medical_form(request):
    if request.method == 'POST':
        form_data = fund_raising_form(
            category=request.POST.get("category"),
            fund_for=request.POST.get("fund_for"),
            fund_title=request.POST.get("fund_title"),
            patienet_full_name=request.POST.get("patienet_full_name"),
            patienet_age=request.POST.get("patienet_age"),
            hospital_name=request.POST.get("hospital_name"),
            hospital_location=request.POST.get("hospital_location"),
            hospitalization_status=request.POST.get("hospitalization_status"),
            fund_raising_amount=request.POST.get("fund_raising_amount"),
            fund_purpose=request.POST.get("fund_purpose"),
            start_date=date.today(),
            end_date=date.today(),
            created_date=date.today(),
            user_id=123,
            user_info="Unknown",
            created_by="system",
        )

        # Convert image file to base64 string
        if request.FILES.get("attachments"):
            image_file = request.FILES.get("attachments")
            image_data = image_file.read()
            encoded_image = base64.b64encode(image_data).decode('utf-8')
            form_data.patienet_image = encoded_image

        form_data.save()
        return redirect('dashboard')

    return render(request, 'medical_form.html')


from datetime import date
from django.shortcuts import render, redirect

def ngo_form(request):
    if request.method == 'POST':
        form_data = fund_raising_form(
            category=request.POST.get("category"),
            fund_title=request.POST.get("fund_title", ""),
            ngo_name=request.POST.get("ngo_name", ""),
            ngo_registration_number=request.POST.get("ngo_registration_number", ""),
            ngo_website_url=request.POST.get("ngo_website_url", ""),
            user_qualification=request.POST.get("user_qualification", ""),
            user_employed_status=request.POST.get("user_employed_status", ""),
            how_to_know=request.POST.get("how_to_know", ""),
            fund_raising_amount=request.POST.get("fund_raising_amount", 0),
            start_date=date.today(),
            end_date=date.today(),
            created_date=date.today(),
            user_id=123,
            user_info="Unknown",
            created_by="system",
        )
        form_data.save()
        return redirect('dashboard')

    return render(request, 'ngo_form.html')

import base64
from datetime import date
from django.shortcuts import render, redirect

def other_form(request):
    if request.method == 'POST':
        form_data = fund_raising_form(
            category=request.POST.get("category"),
            fund_title=request.POST.get("fund_title", ""),
            fund_for=request.POST.get("fund_for", ""),
            fund_raising_amount=request.POST.get("fund_raising_amount", 0),
            user_qualification=request.POST.get("user_qualification", ""),
            user_employed_status=request.POST.get("user_employed_status", ""),
            how_to_know=request.POST.get("how_to_know", ""),
            fund_purpose=request.POST.get("fund_purpose", ""),
            start_date=date.today(),
            end_date=date.today(),
            created_date=date.today(),
            user_id=123,
            user_info="Unknown",
            created_by="system",
        )

        # Convert image file to base64 string if uploaded
        if request.FILES.get("attachments"):
            image_file = request.FILES.get("attachments")
            image_data = image_file.read()
            encoded_image = base64.b64encode(image_data).decode('utf-8')
            form_data.patienet_image = encoded_image

        # Save medical_documents file normally if your model allows it (optional)
        if request.FILES.get("medical_documents"):
            form_data.medical_documents = request.FILES.get("medical_documents")

        form_data.save()
        return redirect('dashboard')

    return render(request, 'other_form.html')

import openpyxl
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage

# def download_students_excel(request):
    
#     # Workbook aur worksheet banao
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Students"
    
#     # Header row
#     headers = ['Name', 'Class', 'Roll Number']
#     ws.append(headers)
    
#     # Data rows from database
#     for student in Student.objects.all():
#         ws.append([student.name, student.student_class, student.roll_number])
        
#     # Response set karo
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    
#     # Excel file ko response me save karo
#     wb.save(response)
#     return response

def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        
        fs = FileSystemStorage()
        
        filename = fs.save(excel_file.name, excel_file)
        
        uploaded_file_url = fs.url(filename)
        
        import pandas as pd
        df = pd.read_excel(fs.path(filename))
        
        return render(request, 'studentinfo.html', {'uploaded_file_url': uploaded_file_url})
    
def download_excel_page(request):
    return render(request, 'download_medicalleads.html')
    
def download_medicalleads_excel(request):
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MedicalLeads"

    # Header row – must match model fields (excluding files/images)
    headers = [
        'lead_generate_by', 'lead_purpose', 'patient_name', 'patient_address', 'patient_age',
        'patient_case', 'claim_amount', 'estimate_amount', 'estimate_date', 'patient_education',
        'patient_occupation', 'guardian_name', 'guardian_number', 'hospital_name',
        'doctor_number', 'case_type', 'lead_source'
    ]
    ws.append(headers)

    # Data rows
    for lead in MedicalLead.objects.all():
        ws.append([
            lead.lead_generate_by,
            lead.lead_purpose,
            lead.patient_name,
            lead.patient_address,
            lead.patient_age,
            lead.patient_case,
            lead.claim_amount,
            lead.estimate_amount,
            lead.estimate_date,
            lead.patient_education,
            lead.patient_occupation,
            lead.guardian_name,
            lead.guardian_number,
            lead.hospital_name,
            lead.doctor_number,
            lead.case_type,
            lead.lead_source
        ])

    # Send Excel as downloadable response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=medicalleads.xlsx'

    wb.save(response)
    return response


from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import MedicalLead
from .serializers import MedicalLeadSerializer 

# GET all + POST new
@api_view(['GET', 'POST'])
def medicallead_list(request):
    if request.method == 'GET':
        leads = MedicalLead.objects.all()
        serializer = MedicalLeadSerializer(leads, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = MedicalLeadSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response (status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# GET one + PUT + DELETE
@api_view(['GET', 'PUT', 'DELETE'])
def medicallead_detail(request, pk):
    try:
        lead = MedicalLead.objects.get(pk=pk)
    except MedicalLead.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = MedicalLeadSerializer(lead)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = MedicalLeadSerializer(lead, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        lead.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)




    


